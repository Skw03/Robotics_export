#!/usr/bin/env python3

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "deliverables" / "robotics_experiment_template.xlsx"
DEFAULT_RESULTS_DIR = ROOT / "experiment_results" / "real_midterm"


EXPERIMENT_HEADERS = [
    "trial",
    "scene",
    "task",
    "planner_profile",
    "avoidance_profile",
    "accepted",
    "status",
    "task_status",
    "elapsed_sec",
    "route",
    "error",
    "notes",
]


def read_csv_records(results_dir: Path) -> List[Dict[str, str]]:
    records: List[Dict[str, str]] = []
    for csv_path in sorted(results_dir.glob("*.csv")):
        with csv_path.open("r", encoding="utf-8", newline="") as stream:
            reader = csv.DictReader(stream)
            for row in reader:
                row = {key: value for key, value in row.items()}
                row["source_file"] = csv_path.name
                records.append(row)
    return records


def normalize_record(record: Dict[str, str]) -> List[object]:
    values: List[object] = []
    for header in EXPERIMENT_HEADERS:
        value = record.get(header, "")
        if header == "trial":
            try:
                values.append(int(value))
            except ValueError:
                values.append(value)
        elif header == "elapsed_sec":
            try:
                values.append(float(value))
            except ValueError:
                values.append(value)
        elif header == "accepted":
            values.append(str(value).lower() in ("true", "1", "yes"))
        else:
            values.append(value)
    return values


def style_header(ws, color: str):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def reset_sheet(ws, headers: Iterable[str], color: str):
    ws.delete_rows(1, ws.max_row)
    ws.append(list(headers))
    style_header(ws, color)
    ws.freeze_panes = "A2"


def write_experiment_sheet(wb, records: List[Dict[str, str]]):
    ws = wb["Experiment Template"]
    reset_sheet(ws, EXPERIMENT_HEADERS, "1F4E78")
    for record in records:
        ws.append(normalize_record(record))

    widths = {
        "A": 10,
        "B": 14,
        "C": 14,
        "D": 18,
        "E": 22,
        "F": 12,
        "G": 16,
        "H": 16,
        "I": 14,
        "J": 52,
        "K": 42,
        "L": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_summary_sheet(wb):
    ws = wb["Summary"]
    ws.delete_rows(1, ws.max_row)
    rows = [
        ["Metric", "Formula / value"],
        ["Total trials", "=COUNTA('Experiment Template'!A2:A200)"],
        ["Completed trials", '=COUNTIF(\'Experiment Template\'!H2:H200,"COMPLETED")'],
        ["Success rate", "=IF(B2=0,0,B3/B2)"],
        ["Average elapsed sec", '=IFERROR(AVERAGEIF(\'Experiment Template\'!H2:H200,"COMPLETED",\'Experiment Template\'!I2:I200),0)'],
        ["Error trials", '=COUNTIF(\'Experiment Template\'!G2:G200,"ERROR")'],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, "548235")
    ws["B4"].number_format = "0.0%"
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 30


def load_llm_json_records(results_dir: Path) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    for json_path in sorted(results_dir.glob("llm_*.json")):
        try:
            records.append(json.loads(json_path.read_text(encoding="utf-8")))
        except json.JSONDecodeError:
            continue
    return records


def write_llm_sheet(wb, results_dir: Path):
    ws = wb["LLM Test Cases"]
    headers = [
        "command",
        "expected_scene",
        "expected_task",
        "parsed_scene",
        "parsed_task",
        "parser",
        "latency_sec",
        "confidence",
        "correct",
        "failure_note",
    ]
    existing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing.append(list(row[: len(headers)]))

    llm_records = load_llm_json_records(results_dir)
    if llm_records:
        existing = []
        for item in llm_records:
            parsed = item.get("parsed", {}) if isinstance(item, dict) else {}
            command = item.get("input", "") if isinstance(item, dict) else ""
            existing.append([
                command,
                "",
                "",
                parsed.get("scene", ""),
                parsed.get("task", ""),
                parsed.get("parser", ""),
                parsed.get("latency_sec", ""),
                parsed.get("confidence", ""),
                "",
                parsed.get("error", ""),
            ])

    reset_sheet(ws, headers, "7030A0")
    for row_idx, row in enumerate(existing, start=2):
        row = list(row) + [""] * (len(headers) - len(row))
        ws.append(row[: len(headers)])
        ws.cell(row_idx, 9).value = f'=IF(AND(B{row_idx}=D{row_idx},C{row_idx}=E{row_idx}),"yes","no")'

    widths = [48, 18, 18, 18, 18, 22, 14, 14, 12, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def parse_args():
    parser = argparse.ArgumentParser(description="Fill the robotics experiment workbook from real midterm CSV logs.")
    parser.add_argument("--results-dir", default=str(DEFAULT_RESULTS_DIR))
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--output", default=str(DEFAULT_WORKBOOK))
    return parser.parse_args()


def main():
    args = parse_args()
    results_dir = Path(args.results_dir)
    workbook_path = Path(args.workbook)
    output_path = Path(args.output)
    records = read_csv_records(results_dir)

    wb = load_workbook(workbook_path)
    write_experiment_sheet(wb, records)
    write_summary_sheet(wb)
    write_llm_sheet(wb, results_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(json.dumps({
        "workbook": str(output_path),
        "records": len(records),
        "results_dir": str(results_dir),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
